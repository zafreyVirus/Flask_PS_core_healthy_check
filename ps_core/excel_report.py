from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Severity colour fills
SEVERITY_FILLS = {
    "Critical": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
    "Major":    PatternFill(start_color="FF8000", end_color="FF8000", fill_type="solid"),
    "Warning":  PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
    "Minor":    PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
}

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

ALARM_COLUMNS = [
    "Severity",
    "Alarm ID",
    "Name",
    "NE Type",
    "Alarm Source",
    "MO Name",
    "Location Information",
    "First Occurred (NT)",
    "Additional Information"
]

ALARM_COL_WIDTHS = [12, 14, 28, 10, 16, 14, 45, 22, 55]


class ExcelReport:

    def __init__(self, author):
        self.author = author

    def create_report(self, traffic_charts, cpu_charts, health_report, output_file, cups_charts=None, usn_alarms=None, ugw_alarms=None):

        wb = Workbook()

        # SHEET 1: PS CORE TRAFFIC

        ws1 = wb.active
        ws1.title = "Weekly PS Core Traffic"

        ws1["A1"] = "PS CORE TRAFFIC REPORT"
        ws1["A2"] = f"Author: {self.author}"

        row = 4

        for chart in traffic_charts:

            img = Image(chart)
            img.width = 900
            img.height = 300

            ws1.add_image(img, f"A{row}")

            row += 22

        # health summary table

        row += 2
        ws1[f"A{row}"] = "NE"
        ws1[f"B{row}"] = "Peak Traffic"
        ws1[f"C{row}"] = "Peak Time"
        ws1[f"D{row}"] = "Min Traffic"
        ws1[f"E{row}"] = "Min Time"
        ws1[f"F{row}"] = "Average"
        ws1[f"G{row}"] = "Utilization %"
        ws1[f"H{row}"] = "Health Status"

        row += 1

        for ne, stats in health_report.items():

            ws1[f"A{row}"] = ne
            ws1[f"B{row}"] = stats["Peak Traffic (MB)"]
            ws1[f"C{row}"] = str(stats["Peak Time"])
            ws1[f"D{row}"] = stats["Minimum Traffic (MB)"]
            ws1[f"E{row}"] = str(stats["Minimum Time"])
            ws1[f"F{row}"] = stats["Average Traffic (MB)"]
            ws1[f"G{row}"] = stats["Utilization %"]
            ws1[f"H{row}"] = stats["Health Status"]

            row += 1

        # SHEET 2: USN CPU

        ws2 = wb.create_sheet("USN CPU Usage")

        ws2["A1"] = "USN CPU USAGE REPORT"
        ws2["A2"] = f"Author: {self.author}"

        row = 4

        for chart in cpu_charts:

            img = Image(chart)
            img.width = 1100
            img.height = 300

            ws2.add_image(img, f"A{row}")

            row += 22

        # SHEET 3: CUPS CPU

        if cups_charts:

            ws3 = wb.create_sheet("CUPS CPU Usage")

            ws3["A1"] = "CUPS CPU USAGE REPORT"
            ws3["A2"] = f"Author: {self.author}"

            row = 4

            for chart in cups_charts:

                img = Image(chart)
                img.width = 1100
                img.height = 300

                ws3.add_image(img, f"A{row}")

                row += 22

        # SHEET 4: USN ALARMS

        if usn_alarms:

            ws4 = wb.create_sheet("USN Alarms")

            ws4["A1"] = "USN ALARMS REPORT"
            ws4["A1"].font = Font(bold=True, size=14)
            ws4["A2"] = f"Author: {self.author}"

            current_row = 4

            for section_title, df in usn_alarms:

                # Section title
                ws4[f"A{current_row}"] = section_title
                ws4[f"A{current_row}"].font = Font(bold=True, size=12, color="1F4E79")
                current_row += 1

                # Column headers
                for col_idx, col_name in enumerate(ALARM_COLUMNS, start=1):
                    cell = ws4.cell(row=current_row, column=col_idx, value=col_name)
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                current_row += 1

                # Data rows
                for _, data_row in df.iterrows():

                    for col_idx, col_name in enumerate(ALARM_COLUMNS, start=1):
                        value = data_row.get(col_name, "")
                        cell = ws4.cell(row=current_row, column=col_idx, value=str(value) if value else "")
                        cell.border = THIN_BORDER
                        cell.alignment = Alignment(vertical="center", wrap_text=True)

                        # Colour the Severity cell
                        if col_name == "Severity":
                            severity = str(value).strip()
                            if severity in SEVERITY_FILLS:
                                cell.fill = SEVERITY_FILLS[severity]
                                cell.font = Font(bold=True, color="FFFFFF")
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                    current_row += 1

                # Gap between tables
                current_row += 2

            # Set column widths
            for col_idx, width in enumerate(ALARM_COL_WIDTHS, start=1):
                ws4.column_dimensions[get_column_letter(col_idx)].width = width

        # SHEET 5: UGW ALARMS

        if ugw_alarms:

            ws5 = wb.create_sheet("UGW Alarms")

            ws5["A1"] = "UGW ALARMS REPORT"
            ws5["A1"].font = Font(bold=True, size=14)
            ws5["A2"] = f"Author: {self.author}"

            current_row = 4

            for section_title, df in ugw_alarms:

                # Section title
                ws5[f"A{current_row}"] = section_title
                ws5[f"A{current_row}"].font = Font(bold=True, size=12, color="1F4E79")
                current_row += 1

                # Column headers
                for col_idx, col_name in enumerate(ALARM_COLUMNS, start=1):
                    cell = ws5.cell(row=current_row, column=col_idx, value=col_name)
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                current_row += 1

                # Data rows
                for _, data_row in df.iterrows():

                    for col_idx, col_name in enumerate(ALARM_COLUMNS, start=1):
                        value = data_row.get(col_name, "")
                        cell = ws5.cell(row=current_row, column=col_idx, value=str(value) if value else "")
                        cell.border = THIN_BORDER
                        cell.alignment = Alignment(vertical="center", wrap_text=True)

                        if col_name == "Severity":
                            severity = str(value).strip()
                            if severity in SEVERITY_FILLS:
                                cell.fill = SEVERITY_FILLS[severity]
                                cell.font = Font(bold=True, color="FFFFFF")
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                    current_row += 1

                # Gap between tables
                current_row += 2

            # Set column widths
            for col_idx, width in enumerate(ALARM_COL_WIDTHS, start=1):
                ws5.column_dimensions[get_column_letter(col_idx)].width = width

        wb.save(output_file)